import csv
import openpyxl
import send2trash
import textwrap
from datetime import datetime
from tkinter import Tk
from tkinter.filedialog import askopenfilename, asksaveasfilename
from openpyxl.styles import Font

# Variables
error_log_reader = []
number_errors = 0
tipps_reader = []


# User defined functions
def magic_snip(row, number_errors):
    # Gets rid of unwanted rows and splits the Name column into title and package columns
    if row['Old'] == row['New']:  # getting rid of rows with no change
        return
    elif row['Property'] == 'coverageDepth' \
            or row['Property'] == 'impId' \
            or row['Property'] == 'payment' \
            or row['Property'] == 'pkg' \
            or row['Property'] == 'platform' \
            or row['Property'] == 'rectype' \
            or row['Property'] == 'status':  # getting rid of rows that we don't need
        return
    else:
        tip = row['Name']
        if tip.count(' / ') == 1:
            title, package = tip.split(' / ')
            tipps_reader.append(
                [row['Timestamp'], title, package, row['Event'], row['Property'], row['Actor'], row['Old'], row['New'],
                 row['Link']])
        elif tip.count(' / ') > 1:
            pit = tip[::-1]
            package_rev, title_rev = pit.split(' / ', 1)
            title = title_rev[::-1]
            package = package_rev[::-1]
            tipps_reader.append(
                [row['Timestamp'], title, package, row['Event'], row['Property'], row['Actor'], row['Old'], row['New'],
                 row['Link']])
        else:
            error_log_reader.append(
                [row['Timestamp'], row['Name'], row['Event'], row['Property'], row['Actor'], row['Old'], row['New'],
                 row['Link']])
            number_errors = number_errors + 1
    return


# Start of program stuff

print(textwrap.fill('This program will format a DM Change log to be sent to Ex Libris.'))
print()
print(textwrap.fill('You will need to run a DM change log report with the following options:'))
print(textwrap.fill(' - The correct dates'))
print(textwrap.fill(' - ALL (Real Users)'))
print(textwrap.fill(' - TIPPS'))
print(textwrap.fill(' - New Items and Updates to existing items'))
print()
print(textwrap.fill('Click this window then press enter or return to open your file. You can press Ctrl-C at any time '
                    'to quit.'))
input()

# Open original file

Tk().withdraw()
originalFilename = askopenfilename(initialdir="C:\\")

##Use for testing
##os.chdir(r'C:\Users\gemma.wright\Jisc\OneDrive - Jisc\To -do\DM change log')
##originalFilename = 'original.csv'


errorLogFilename = 'errorLog ' + datetime.now().strftime('%Y-%m-%d %H%M%S') + '.csv'
tempFilename = 'temp' + datetime.now().strftime('%Y-%m-%d %H%M%S') + '.csv'

# Get rid of the useless rows, split the name into title and package columns.

with open(originalFilename, encoding='utf-8') as originalFile:
    for i in range(2):
        next(originalFile)

    originalCSV = csv.DictReader(originalFile)

    tipps_reader.append(['Timestamp', 'Title', 'Package', 'Event', 'Property', 'Actor', 'Old', 'New', 'Link'])

    # magic snip

    for row in originalCSV:
        if row['Event'] == 'New TIPP' or row['Event'] == 'Updated TIPP':
            magic_snip(row, number_errors)

        else:
            error_log_reader.append([row['Timestamp'], row['Name'], row['Event'], row['Property'], row['Actor'],
                                     row['Old'], row['New'], row['Link']])
            number_errors = number_errors + 1

# Save to a temp file

with open(tempFilename, 'w', encoding='utf-8', newline='') as TippsFile:
    TippsWriter = csv.writer(TippsFile)
    for row in tipps_reader:
        TippsWriter.writerow(row)

# Create error log if there are errors

if number_errors > 0:
    with open(errorLogFilename, 'w', encoding='utf-8', newline='') as errorLogFile:
        errorLogWriter = csv.writer(errorLogFile)
        for row in error_log_reader:
            errorLogWriter.writerow(row)

# Create the Excel workbook

with open(tempFilename, encoding='utf-8') as TippsFile:
    TippsCSV = csv.DictReader(TippsFile)
    workbook = openpyxl.Workbook()
    packageList = {}

    for row in TippsCSV:
        # Do some variables:

        packageName = row['Package']
        shortName = (
            packageName[:25] + '...' if len(packageName) > 25 else packageName)  # truncate so it doesn't break Excel

        # Get rid of non-Jisc packages

        if 'jisc' not in packageName.lower():
            continue

        # check if package is already a sheet
        # if it is, add the row to that sheet
        elif packageName in packageList:
            # add data to end of correct sheet
            sheet = workbook[packageList[packageName]]
            sheet.append([row['Timestamp'], row['Title'], row['Package'], row['Event'], row['Property'], row['Old'],
                          row['New']])

        # if it isn't, add a new sheet and add the row to the new sheet
        else:
            # create sheet

            for i in range(1, len(packageList) + 2):
                sheetName = shortName.replace(':', '_') + ' ' + str(i)  # replace : with _ to not break Excel

                if sheetName in packageList.values():
                    continue
                else:
                    sheet = workbook.create_sheet(title=sheetName)
                    # add to packageList
                    packageList[packageName] = sheetName
                    # add header
                    sheet['A1'] = packageName
                    # add column headings
                    sheet.append(['Timestamp', 'Title', 'Package', 'Event', 'Property', 'Old', 'New'])
                    # add data to row 2
                    sheet.append(
                        [row['Timestamp'], row['Title'], row['Package'], row['Event'], row['Property'], row['Old'],
                         row['New']])
                    break

# Create summary page
# Rename 'Sheet'
workbook['Sheet'].title = 'Summary'

with open(originalFilename, encoding='utf-8') as originalFile:
    originalHeader = csv.reader(originalFile)
    headerList = list(originalHeader)
    startDate = headerList[1][0].strip().replace(r'"', '')
    endDate = headerList[1][1].strip().replace(r'"', '')

sheet = workbook['Summary']
sheet['A1'] = 'Changes to Jisc Collections packages'
sheet.append(['Start date:', startDate])
sheet.append(['End date:', endDate])

sheet['A5'] = 'Packages changed:'
sheet['B5'] = 'Sheet name:'

for i in packageList:
    sheet.append([i, packageList[i]])

# Prettify the worksheets
fontBold = Font(bold=True)
fontNotBold = Font(bold=False)

for sheet in workbook.worksheets:
    sheet['A1'].font = fontBold
    for i in range(1, 8):
        sheet.cell(row=2, column=i).font = fontBold
    sheet.column_dimensions['A'].width = 19.14
    sheet.column_dimensions['B'].width = 45
    sheet.column_dimensions['C'].width = 45
    sheet.column_dimensions['D'].width = 12.14
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 30

# Prettify summary

sheet = workbook['Summary']

sheet['A2'].font = fontNotBold
sheet['B2'].font = fontNotBold
sheet['A5'].font = fontBold
sheet['B5'].font = fontBold

sheet.column_dimensions['A'].width = 85
sheet.column_dimensions['B'].width = 30

# End of program stuff

print(textwrap.fill('The change log is ready to save. There were ' + str(number_errors) + ' error(s).'))
if number_errors > 0:
    print('Errors are record in the errors log: ' + str(errorLogFilename))

print()
print(textwrap.fill('Press return to save the file.'))
input()

outputFilename = asksaveasfilename(initialfile='JC changes ' + startDate + ' to ' + endDate,
                                   defaultextension=".xlsx")

##Use for testing
##outputFilename = 'output' + datetime.now().strftime('%Y-%m-%d %H%M%S')  + '.xlsx'            


# Save the workbook

workbook.save(outputFilename)
send2trash.send2trash(tempFilename)
